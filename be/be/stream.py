import asyncio
from datetime import datetime
import logging

import pandas as pd
import aiohttp
from openpyxl import load_workbook

from .types import Comment, Score, ScoreComment
from .cache import acache


LOGGER = logging.getLogger(__name__)
API_URL = "http://localhost:10080/msg"


def setup(production: bool):
    global API_URL
    if production:
        API_URL = "https://text-score-hvi4fczdoa-an.a.run.app/msg"
    LOGGER.info("stream setup: %s", API_URL)


def read_data0() -> pd.DataFrame:
    df = (
        pd
        .read_csv("data/output.csv")
        .query("race_name == 'ル・マン24時間フリー走行・予選'")
        .assign(
            created_at=lambda df: [
                datetime.strptime(x, "%Y-%m-%d %H:%M:%S+00:00")
                for x in df["created_at"]
            ]
        )
        .sort_values("created_at")
        .assign(
            diff=lambda df: [
                x.seconds
                for x in (df["created_at"] - df["created_at"].iloc[0])
            ]
        )
    )
    return df


def read_data1() -> pd.DataFrame:
    wb = load_workbook("data/20230803-yt_chat_annotation.xlsx")
    ws = wb.worksheets[0]
    n = ws.max_row
    time = [
        ws.cell(i, 2).value
        for i in range(2, n + 1)
    ]
    text = [
        ws.cell(i, 5).value
        for i in range(2, n + 1)
    ]
    return (
        pd
        .DataFrame({
            "created_at": time,
            "text": text
        })
        .sort_values("created_at")
        .assign(
            created_at=lambda df: [
                datetime.strptime(x, "%Y-%m-%d %H:%M:%S")
                for x in df["created_at"]
            ],
            diff=lambda df: [
                x.seconds
                for x in (df["created_at"] - df["created_at"].iloc[0])
            ]
        )
    )


async def read_data(text_queue: asyncio.Queue[Comment]):
    # df = read_data0()
    df = read_data1()
    last_diff = 0
    for i, row in df.iterrows():
        obj = Comment(time=row["created_at"], text=row["text"])  # type: ignore
        await text_queue.put(obj)
        LOGGER.info("put %s", obj)
        diff = row["diff"] - last_diff
        LOGGER.info("sleeping %s", diff)
        # await asyncio.sleep(diff)
        await asyncio.sleep(1)
        last_diff = row["diff"]


async def get_score(text_queue: asyncio.Queue[Comment], score_queue: asyncio.Queue[ScoreComment]):
    while True:
        comment = await text_queue.get()
        LOGGER.info("get text %s", comment)
        vals = await score(comment.text)
        LOGGER.info("receiving score %s", vals)
        cs = ScoreComment(score=vals, comment=comment)
        await score_queue.put(cs)


@acache
async def score(text: str) -> Score:
    async with aiohttp.ClientSession() as session:
        payload = {
            "message": text
        }
        async with session.post(API_URL, json=payload) as resp:
            ret = await resp.json()
            return Score.model_validate(ret["scores"])
